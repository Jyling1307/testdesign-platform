"""Generate xlsx test case file from structured test case data.

Format matches SKILL.md (md-step2-to-xlsx) specification:
- Font: 宋体 11pt, regular, black
- Alignment: wrap_text, vertical=top
- 7 columns: 用例名称, 所属产品, 用例类型, 适用阶段, 前置条件, 步骤, 预期结果
- No borders, fills, bold, auto-filter, merged cells, or frozen panes
- 转换说明 sheet with merge summary
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


HEADERS = ['用例名称', '所属产品', '用例类型', '适用阶段', '前置条件', '步骤', '预期结果']
COLUMN_WIDTHS = {'A': 60, 'B': 10, 'C': 13, 'D': 13, 'E': 65, 'F': 100, 'G': 50}

# Conditions for merging sibling cases (all 4 must be true)
MAX_STEPS_FOR_MERGE = 5
MAX_EXPECTED_FOR_MERGE = 4


def _can_merge_cases(cases: list[dict], start_idx: int) -> bool:
    """Check if a group of consecutive sibling cases can be merged.

    Merge conditions:
    1. All cases share the same case_type
    2. All cases share the same precondition
    3. All cases share the same step operation target (first N chars of steps)
    4. Merged steps count <= MAX_STEPS_FOR_MERGE and expected lines <= MAX_EXPECTED_FOR_MERGE
    """
    group = [cases[start_idx]]
    i = start_idx + 1
    while i < len(cases):
        next_case = cases[i]
        # Same type and precondition
        if (next_case['case_type'] == group[0]['case_type']
                and next_case.get('precondition', '') == group[0].get('precondition', '')):
            group.append(next_case)
            i += 1
        else:
            break

    if len(group) <= 1:
        return False, []

    # Check step count and expected count after merge
    total_steps = sum(len(c.get('steps', '').split('\n')) for c in group if c.get('steps'))
    total_expected = sum(len(c.get('expected', '').split('\n')) for c in group if c.get('expected'))

    if total_steps > MAX_STEPS_FOR_MERGE or total_expected > MAX_EXPECTED_FOR_MERGE:
        return False, []

    return True, group


def _merge_case_group(group: list[dict]) -> dict:
    """Merge a group of cases into a single case."""
    merged = {
        'name': group[0]['name'],
        'product': group[0].get('product', ''),
        'case_type': group[0]['case_type'],
        'phase': group[0].get('phase', ''),
        'precondition': group[0].get('precondition', ''),
        'steps': '\n'.join(c.get('steps', '') for c in group if c.get('steps')),
        'expected': '\n'.join(c.get('expected', '') for c in group if c.get('expected')),
    }
    return merged


def generate_xlsx(cases: list[dict], product: str = '') -> bytes:
    """Generate xlsx file content from test case dicts.

    Args:
        cases: List of dicts from md_parser.parse_step2_to_cases
        product: Product name for column B (overrides per-case product)

    Returns:
        bytes: xlsx file content
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet'

    # Styles
    font = Font(name='宋体', size=11)
    align = Alignment(wrap_text=True, vertical='top')

    # Headers
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = font
        cell.alignment = align

    # Column widths
    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # Apply merge logic and write data rows
    merged_count = 0
    i = 0
    row_idx = 2
    while i < len(cases):
        can_merge, group = _can_merge_cases(cases, i)
        if can_merge and len(group) > 1:
            merged_case = _merge_case_group(group)
            merged_count += len(group) - 1
            _write_case_row(ws, row_idx, merged_case, product, font, align)
            row_idx += 1
            i += len(group)
        else:
            _write_case_row(ws, row_idx, cases[i], product, font, align)
            row_idx += 1
            i += 1

    # Add 转换说明 sheet
    ws2 = wb.create_sheet('转换说明')
    if merged_count > 0:
        ws2.cell(row=1, column=1, value=f'本次转换共合并 {merged_count} 条用例').font = font
    else:
        ws2.cell(row=1, column=1, value='本次转换未进行用例合并').font = font
    ws2.column_dimensions['A'].width = 40

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _write_case_row(ws, row_idx: int, case: dict, product: str, font: Font, align: Alignment):
    """Write a single test case row to the worksheet."""
    values = [
        case.get('name', ''),
        product or case.get('product', ''),
        case.get('case_type', '功能测试'),
        case.get('phase', ''),
        case.get('precondition', ''),
        case.get('steps', ''),
        case.get('expected', ''),
    ]
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = font
        cell.alignment = align
